"""
# main.py
import pandas as pd

# === 1. Cargar la base de datos ===
file_path = "DATA/expo-julio25.xlsx"  # cambia por la ruta real
df = pd.read_excel(file_path)
print(df)
"""
"""
# Verificar estructura
print(df.head())

# === 2. Limpieza básica ===
# Estandarizar nombres de columnas a minúsculas
df.columns = [c.strip().lower() for c in df.columns]

# Convertir columnas clave
df['año'] = df['año'].astype(int)
df['flujo'] = df['flujo'].str.upper()

# === 3. Tabla resumen total ===
resumen_total = df.groupby(['flujo', 'año']).size().reset_index(name='n_operaciones')

# Si tienes columna de valor FOB, podrías usar sum()
# resumen_total = df.groupby(['flujo', 'año'])['valor_fob'].sum().reset_index()

print("\n--- Resumen total ---")
print(resumen_total)

# === 4. Exportar a Excel para usar en el reporte ===
with pd.ExcelWriter("output/resumen_comercio.xlsx") as writer:
    resumen_total.to_excel(writer, sheet_name="Resumen", index=False)
"""

"""
import pandas as pd
import numpy as np
import statsmodels.tsa.stattools as ts
import openpyxl
from tabulate import tabulate

excel_dataframe= openpyxl.load_workbook("expo-julio25.xlsx")

dataframe=excel_dataframe.active

data=[]

for row in range(1, dataframe.max_row): #recuento de los numeros de filas 
    print(row)
    """
"""
    _row = [row,]
    for col in dataframe.iter_cols(1,dataframe.max_column):
        _row.append(col[row].value)

    data.append(_row)
    
headers = ["S&P", "PBI", "TCRM", "TIR", "IPC", "Empleo"] 
headers_align = (("center",)*6)

print(tabulate(data,headers=headers,tablefmt="fancy_grid", colalign=headers_align))
"""


import pandas as pd
from tabulate import tabulate

# 1. Leer el Excel
df = pd.read_excel("DATA/expo-julio25.xlsx")

# 2. Estandarizar columnas
df.columns = [c.strip().lower() for c in df.columns]

# 3. Mostrar las primeras filas como tabla
print("\n📄 Primeras filas de la base:")
print(tabulate(df.head(10), headers=df.columns, tablefmt="fancy_grid", showindex=False))

# 4. Resumen simple por flujo y año
if 'flujo' in df.columns and 'año' in df.columns:
    resumen = df.groupby(['flujo', 'año']).size().reset_index(name='n_operaciones')
    print("\n📊 Resumen por flujo y año:")
    print(tabulate(resumen, headers=resumen.columns, tablefmt="fancy_grid", showindex=False))
else:
    print("⚠️ La base no tiene columnas 'flujo' y/o 'año'.")
