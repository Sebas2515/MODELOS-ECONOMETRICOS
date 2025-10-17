import pandas as pd       # Manipulación y análisis de datos en tablas (DataFrames).
import numpy as np        # Cálculos numéricos y manejo de arreglos/matrices.
import seaborn as sns     # Visualización estadística avanzada (gráficos bonitos y rápidos).
import statsmodels.api as sm       # Modelos econométricos y estadísticos (regresiones, pruebas, etc.).
import statsmodels.stats.api as sms  # Pruebas estadísticas específicas (heterocedasticidad, autocorrelación...).
from statsmodels.stats.outliers_influence import variance_inflation_factor  
# Calcula el VIF (factor de inflación de la varianza) para detectar multicolinealidad.
import matplotlib.pyplot as plt    # Creación de gráficos básicos y personalizables.
from pathlib import Path           # Manejo de rutas y archivos de forma más segura y moderna.
from tabulate import tabulate      # Muestra tablas en consola con formato legible.

### Ver todas las hojas de excel ###
"""
excel = pd.ExcelFile('DATA/model_program.xlsx')
print(excel.sheet_names)  # Lista todas las hojas
"""
#### Ver que todas las hojas de excel y en cual se encuentra activa ###
"""
from openpyxl import load_workbook (modulo Workbook)
wb = load_workbook('DATA/model_program.xlsx')
print(wb.sheetnames)      # Muestra todas las hojas
print(wb.active.title)    # Muestra el nombre de la hoja activa
"""
################################################################################
# PASO 0: CONFIGURACIÓN Y CARGA DE DATOS
################################################################################

path = Path('DATA/model_program.xlsx')

# Cargar la hoja específica para la tesis
df = pd.read_excel(path, sheet_name='bd-tri', index_col=None)

# Limpiar nombres de columnas (buena práctica)
df.columns = df.columns.str.strip().str.replace(' ', '_')

# ✅ Si la columna 'Año' está como índice, traerla de vuelta
if 'Año' not in df.columns and df.index.name == 'Año':
    df.reset_index(inplace=True)

# Renombrar columnas para trabajar más fácil
df = df.rename(columns={
    'Ingresos_Fiscales': 'Ingfisca',
})

print("--- 1. Datos Cargados y Preparados ---")
print(df.head())
print("\nInformación del DataFrame:")
df.info()

################################################################################
# PASO 1: Convertirlo la serie en logaritmos 
################################################################################

# Aplicar logaritmo natural (ln) a las variables positivas
df['ln_PBI'] = np.log(df['PBI'])
df['ln_Ingfisca'] = np.log(df['Ingfisca'])
df['ln_TIR'] = np.log(df['TIR'])
df['ln_TE'] = np.log(df['TE'])
df['ln_EP'] = np.log(df['EP'])

# Crear las diferencias logarítmicas (crecimientos porcentuales aproximados)
df['dln_PBI'] = df['ln_PBI'].diff()
df['dln_Ingfisca'] = df['ln_Ingfisca'].diff()
df['dln_TIR'] = df['ln_TIR'].diff()
df['dln_TE'] = df['ln_TE'].diff()
df['dln_EP'] = df['ln_EP'].diff()

# Crear variable de periodo trimestral
df['Fecha'] = pd.PeriodIndex(df['Año'], freq='Q')

# ✅ Crear dummy para quiebre estructural (ej. 2020Q3)
df['dummy_quiebre'] = (df['Fecha'] >= '2020Q3').astype(int)

# ✅ (Opcional) Crear interacciones con las diferencias logarítmicas
df['dummy_dln_TIR'] = df['dummy_quiebre'] * df['dln_TIR']
df['dummy_dln_Ingfisca'] = df['dummy_quiebre'] * df['dln_Ingfisca']
df['dummy_dln_TE'] = df['dummy_quiebre'] * df['dln_TE']
df['dummy_dln_EP'] = df['dummy_quiebre'] * df['dln_EP']

# Eliminar los primeros NaN generados por la diferencia
df = df.dropna()

# Verificar
print(df[['Año', 'Fecha', 'dummy_quiebre']].tail(10))
print(df['dummy_quiebre'].value_counts())


"""
# Eliminar los primeros NaN generados por la diferencia
df = df.dropna()

# Creamos un nuevo DataFrame solo con las variables de interés (en logaritmos)
df_log = df[['ln_PBI_log1', 'ln_Ingfisca_lag1', 'ln_TIR_lag1', 'ln_TE_lag1']].copy()
print(df_log.head())

print("\n--- 1. Datos transformados a diferencias logarítmicas ---")
print(df[['dln_PBI', 'dln_Ingfisca', 'dln_TIR', 'dln_TE']].head())
"""

################################################################################
# PASO 2: TEST DE ESTACIONARIEDAD (DICKEY-FULLER AUMENTADO)
################################################################################
from statsmodels.tsa.stattools import adfuller

def adf_test(series, name=''):
    "Realiza el test de Dickey-Fuller Aumentado en una serie temporal."
    result = adfuller(series.dropna())
    print(f'\n--- Test de Estacionariedad para: {name} ---')
    print(f'ADF Statistic: {result[0]:.4f}')
    print(f'p-value: {result[1]:.4f}')
    if result[1] <= 0.05:
        print("✅ La serie es estacionaria.")
    else:
        print("❌ La serie no es estacionaria (tiene raíz unitaria).")

# Filtrar columnas con 'ln_'
cols_log = [col for col in df.columns if 'ln_' in col]

print("\n--- 2. Verificando Estacionariedad de las series logarítmicas ---")
for name in cols_log:
    adf_test(df[name], name=name)

###############################################################################
# PASO 3: AJUSTE DEL MODELO MCO
################################################################################
"""
print(df.columns.tolist())
df['Fecha'] = pd.PeriodIndex(df['Año'], freq='Q')

# Crear dummy para quiebre estructural en 2020Q3
df['dummy_quiebre'] = (df['Fecha'] >= '2020Q3').astype(int)

print(df[['Año', 'Fecha', 'dummy_quiebre']].tail(10))
print(df['dummy_quiebre'].value_counts())
"""
"""
print(df['Fecha'].head())
print(df.dtypes)
"""
# Definir variables explicativas y dependiente
Y = df['dln_PBI']
X = df[['dln_TIR', 'dln_Ingfisca', 'dln_TE', 'dln_EP', 
        'dummy_quiebre', 'dummy_dln_TIR', 'dummy_dln_Ingfisca', 'dummy_dln_TE', 'dummy_dln_EP']]
X = sm.add_constant(X)

# 4️⃣ Ajustar modelo MCO simple

model = sm.OLS(Y, X).fit()
residuos = model.resid
print("\n=== RESULTADOS DEL MODELO MCO (Δln variables) ===")
print(model.summary())


# 6️⃣ Crear tabla de coeficientes con intervalos de confianza
coef_table = pd.DataFrame({
    'Coeficiente': model.params,
    'Error Std': model.bse,
    't-stat': model.tvalues,
    'p-value': model.pvalues,
    'IC 0.025': model.conf_int()[0],
    'IC 0.975': model.conf_int()[1]
})

# 7️⃣ Crear tabla resumen general del modelo
summary_table = pd.DataFrame({
    'Estadístico': ['R-squared', 'Adj. R-squared', 'F-statistic', 'Prob (F-statistic)',
                    'No. Observations', 'Log-Likelihood', 'AIC', 'BIC', 'Df Residuals', 'Df Model', 'Covariance Type'],
    'Valor': [model.rsquared, model.rsquared_adj, model.fvalue, model.f_pvalue,
              int(model.nobs), model.llf, model.aic, model.bic, model.df_resid, model.df_model, 'nonrobust']
})

# 8️⃣ Mostrar todo con tabulate
print("\n=== Resumen General del Modelo ===")
print(tabulate(summary_table, headers='keys', tablefmt='fancy_grid', floatfmt=".4f"))

print("\n=== Coeficientes del Modelo ===")
print(tabulate(coef_table, headers='keys', tablefmt='fancy_grid', floatfmt=".6f"))


###############################################################################
# PASO 4: PRUEBA DE CORRELACIÓN
################################################################################

# 5️⃣ Seleccionar variables que quieres correlacionar
cols = ['dln_PBI', 'dln_TIR', 'dln_Ingfisca', 'dln_TE']

# 6️⃣ Calcular matriz de correlaciones
corr_matrix = df[cols].corr()

# 7️⃣ Mostrar matriz en consola
print("\n=== Matriz de Correlaciones ===")
print(corr_matrix.round(3))

# 8️⃣ Visualizar matriz con heatmap
plt.figure(figsize=(8,6))
sns.heatmap(corr_matrix, annot=True, cmap='coolwarm', fmt=".2f")
plt.title('Heatmap de Correlaciones')
plt.show()
.0

################################################################################
# PASO 5: PRUEBA DE MULTICOLINEALIDAD VIF
################################################################################

#Calcular VIF
vif_data = pd.DataFrame()
vif_data['Variable'] = X.columns
vif_data['VIF'] = [variance_inflation_factor(X.values, i) for i in range(X.shape[1])]

print(vif_data)


########################## PRUEBA DE AUTOCORRELACION (DURBIN_WATSON) ##########################

from statsmodels.stats.stattools import durbin_watson

dw = durbin_watson(model.resid)
dw_table = pd.DataFrame({
    'Estadístico': ['Durbin-Watson'],
    'Valor': [dw]
})
print("\n=== AUTOCORRELACION _DURBIN_WATSON ===")
print(tabulate(dw_table, headers='keys', tablefmt='fancy_grid', floatfmt=".4f"))

if dw> 2.0: 
    print("Posible autocorrelación negativa")
elif dw < 2.0:
    print("Posible autocorrelación positiva")
elif dw == 2.0:
    print("No hay autocorrelación (ideal)")


# Interpretación:
# DW ≈ 2.0: No hay autocorrelación (ideal).
# DW < 2.0: Posible autocorrelación positiva.
# DW > 2.0: Posible autocorrelación negativa.

################################################################################
# PRUEBA DE HETEROCEDASTICIDAD (TEST DE WHITE)
################################################################################

from statsmodels.stats.diagnostic import het_white
# Test de White
white_test = het_white(model.resid, model.model.exog)

white_test_table = pd.DataFrame({
    'Estadístico': ['LM stat', 'LM p-value', 'F-stat', 'F p-value'],
    'Valor': white_test
})

print("\n=== Resultado del Test de White (Heterocedasticidad) ===")
print(tabulate(white_test_table, headers='keys', tablefmt='fancy_grid', floatfmt=".4f"))

if white_test[1] > 0.05:
    print("No se rechaza H0: la varianza de los errores es constante (homocedástica)")
else:
    print("Se rechaza H0: la varianza de los errores no es constante (heterocedástica)")

# H0 (Hipótesis nula): La varianza de los errores es constante (homocedasticidad)
# H1 (Hipótesis alternativa): La varianza de los errores no es constante (heterocedasticidad)

# Nota: Si el p-value del test es mayor que el nivel de significancia (por ejemplo, 0.05),
# no se rechaza H0, lo que indica que no hay evidencia significativa de heterocedasticidad"""


################################################################################
# PRUEBA DE NORMALIDAD EN LOS RESIDUOS (JARQUE - BERA)
################################################################################

from statsmodels.stats.stattools import jarque_bera

jb_stat, jb_pvalue, skew, kurtosis = jarque_bera(residuos)

jb_table = pd.DataFrame({
    'Estadístico': ['JB estadístico', 'p-value', 'Skew', 'Kurtosis'],
    'Valor': [jb_stat, jb_pvalue, skew, kurtosis]
})

print("\n=== Jarque - Bera ===")
print(tabulate(jb_table, headers='keys', tablefmt='fancy_grid', floatfmt=".4f"))


print("Jarque-Bera Test")
print("JB estadístico:", jb_stat)
print("p-value:", jb_pvalue)
print("Skew:", skew)
print("Kurtosis:", kurtosis)


if jb_pvalue > 0.05:
    print("No se rechaza H0: los residuos se distribuyen normalmente")
else:
    print("Se rechaza H0: los residuos no son normales")


################################################################################
# FUNCIÓN PARA EL TEST DE CHOW - ESTABILIDAD ESTRUCTURAL 
################################################################################
from scipy import stats

def chow_test(df, split_index):
    """Realiza el test de Chow en el punto de quiebre indicado (por posición)"""
    
    # Dividir usando posición, no etiquetas (iloc)
    Y1 = df.iloc[:split_index]['dln_PBI']
    X1 = sm.add_constant(df.iloc[:split_index][['dln_TIR', 'dln_Ingfisca', 'dln_TE', 'dln_EP',
        'dummy_quiebre', 'dummy_dln_TIR', 'dummy_dln_Ingfisca', 'dummy_dln_TE', 'dummy_dln_EP']])
    
    Y2 = df.iloc[split_index:]['dln_PBI']
    X2 = sm.add_constant(df.iloc[split_index:][['dln_TIR', 'dln_Ingfisca', 'dln_TE', 'dln_EP',
        'dummy_quiebre', 'dummy_dln_TIR', 'dummy_dln_Ingfisca', 'dummy_dln_TE', 'dummy_dln_EP']])
    
    # Modelo completo
    Y_full = df['dln_PBI']
    X_full = sm.add_constant(df[[ 'dln_TIR', 'dln_Ingfisca', 'dln_TE', 'dln_EP',
        'dummy_quiebre', 'dummy_dln_TIR', 'dummy_dln_Ingfisca', 'dummy_dln_TE', 'dummy_dln_EP']])
    
    model_full = sm.OLS(Y_full, X_full).fit()
    model1 = sm.OLS(Y1, X1).fit()
    model2 = sm.OLS(Y2, X2).fit()
    
    # Estadístico F de Chow
    n1, n2 = len(Y1), len(Y2)
    k = X_full.shape[1]
    SSR_full = sum(model_full.resid ** 2)
    SSR1 = sum(model1.resid ** 2)
    SSR2 = sum(model2.resid ** 2)
    
    F = ((SSR_full - (SSR1 + SSR2)) / k) / ((SSR1 + SSR2) / (n1 + n2 - 2 * k))
    p_value = 1 - stats.f.cdf(F, k, n1 + n2 - 2 * k)
    return F, p_value

################################################################################
# EVALUAR TODOS LOS POSIBLES PUNTOS DE QUIEBRE
################################################################################

# Reiniciamos el índice para que el loop funcione bien (Año pasa a columna normal)
df_reset = df.reset_index(drop=False).rename(columns={'index': 'Trimestre'})

results = []
for i in range(8, len(df_reset) - 8):  # evita cortes con pocas observaciones
    F, p = chow_test(df_reset, i)
    results.append((df_reset.loc[i, 'Año'], F, p))

results_df = pd.DataFrame(results, columns=['Trimestre', 'F_stat', 'p_value'])

################################################################################
# MOSTRAR RESULTADOS
################################################################################

best_break = results_df.loc[results_df['F_stat'].idxmax()]

print("📊 Test de Chow — Resultados por Trimestre")
print(results_df.to_string(index=False))
print("\n🏆 Posible punto de cambio estructural:")
print(best_break)

if best_break['p_value'] < 0.05:
    print(f"\n❌ Se rechaza H₀: cambio estructural detectado en el trimestre {best_break['Trimestre']}")
else:
    print(f"\n✅ No se rechaza H₀: el modelo es estable estructuralmente")



# prueba: H0: dln_EP + dummy_dln_EP = 0
t_test = model.t_test("dln_EP + dummy_dln_EP = 0")
print(t_test)
